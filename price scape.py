import urllib3
from bs4 import BeautifulSoup
import openpyxl

# scrapes the retail price of switch games

file = 'Prices.xlsx'
wb = openpyxl.load_workbook(filename=file)
ws = wb['Sheet1']

http = urllib3.PoolManager()
qoute_page = 'https://eshop-prices.com/prices?currency=USD'
response = http.request('GET', qoute_page)

# parse the html using beautiful soup and store in variable `soup`
soup = BeautifulSoup(response.data, 'html.parser')

num = 2
for list in soup.body.find_all('tr'):
    title = list.find('th')
    name = title.find('a')
    prices = list.find_all('td')
    if None != name:
        ws.cell(row=num, column=1, value=name.text)
        print(name.text)
    if len(prices) > 0:
        print(prices[len(prices) - 1].text)
        price = prices[len(prices) - 1].text
        if price == 'N/A':
            ws.cell(row=num, column=2, value=price)
        else:
            ws.cell(row=num, column=2, value=float(price[1:]))
    num += 1
wb.save(file)
