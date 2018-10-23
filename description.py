import ebaysdk
import ebaysdk.shopping
from ebaysdk.shopping import Connection as shopping
from ebaysdk.exception import ConnectionError
import openpyxl

api = shopping(siteid='EBAY-US', appid='KevinDeC-deals-PRD-d66850dff-22c77c1b', config_file=None)

"""execl file"""
file = 'VideoGames.xlsx'
wb = openpyxl.load_workbook(filename=file)
ws = wb['Sheet1']

cur = ws['H2'].value
Ids = [cur]
row = 1
num = 2
while cur != None:
    if len(Ids) == 20:
        api.execute('GetMultipleItems', {
            'ItemID': Ids,
            'IncludeSelector': ['ItemSpecifics']
        })
        dictstr = api.response.dict()
        for item in dictstr['Item']:
            row += 1
            if 'ItemSpecifics' in item:
                for spec in item['ItemSpecifics']['NameValueList']:
                    if type(spec) == dict:
                        if 'Game Name' == spec['Name']:
                            ws.cell(row=row, column=9, value='{}'.format(spec['Value']))
                        if 'Publisher' == spec['Name']:
                            ws.cell(row=row, column=10, value='{}'.format(spec['Value']))
        Ids.clear()
    num += 1
    cur = ws['H' + str(num)].value
    Ids.append(cur)
wb.save(file)