import ebaysdk
from ebaysdk.finding import Connection as finding
from ebaysdk.exception import ConnectionError
import openpyxl

# api = finding(domain='svcs.sandbox.ebay.com', appid='KevinDeC-deals-SBX-72ccbdebc-3b9bdca0', config_file=None) // test on ebay's sandbox site
api = finding(siteid='EBAY-US', appid='KevinDeC-deals-PRD-d66850dff-22c77c1b', config_file=None)

api.execute('findItemsAdvanced', {
    'categoryId': '139973',  # Video Games
    'itemFilter': [
        {'name': 'Condition', 'value': 'New'},
        {'name': 'HideDuplicateItems', 'value': True},
        {'name': 'ListingType', 'value': 'StoreInventory'}
    ],
    'aspectFilter': [
        {'aspectName': 'Release Year', 'aspectValueName': ['2018', '2017']},
        {'aspectName': 'Platform', 'aspectValueName': 'Nintendo Switch'}
    ],
    'paginationInput': {
        'entriesPerPage': '100',
        'pageNumber': '1'
    },
    'sortOrder': 'StartTimeNewest',
    'outputSelector': ['GalleryInfo',
                       'PictureURLLarge',
                       'ItemSpecifics'],
    'includeSelector' : 'ItemSpecifics'
})

dictstr = api.response.dict()
# api.execute('getHistograms', {
#         'categoryId': '139973'
# })
# hist = api.response.dict()

"""execl file"""
file = 'VideoGames.xlsx'
wb = openpyxl.load_workbook(filename=file)
ws = wb['Sheet1']

num = 2
for item in dictstr['searchResult']['item']:
    # print(num)

    ws.cell(row=num, column=1, value='{}'.format(item['title']))
    ws.cell(row=num, column=2, value=float('{}'.format(item['sellingStatus']['currentPrice']['value'])))
    ws.cell(row=num, column=3, value='Nintendo')
    ws.cell(row=num, column=4, value='{}'.format(item['primaryCategory']['categoryName']))
    if 'productId' in item:
        ws.cell(row=num, column=5, value=float('{}'.format(item['productId']['value'])))
    ws.cell(row=num, column=6, value='{}'.format(item['viewItemURL']))
    if 'galleryURL' in item:
        ws.cell(row=num, column=7, value='{}'.format(item['galleryURL']))
    ws.cell(row=num, column=8, value='{}'.format(item['itemId']))
    # print('ItemID: {}'.format(item['itemId']))
    # print('Title: {}'.format(item['title']))
    # print('CategoryID: {}'.format(item['primaryCategory']['categoryId']))
    num += 1

for i in range(int(dictstr['paginationOutput']['totalPages'])-1):
    c = i + 1
    api.execute('findItemsAdvanced', {
        'categoryId': '139973',  # Video Games
        'itemFilter': [
            {'name': 'Condition', 'value': 'New'},
            {'name': 'HideDuplicateItems', 'value': True},
            {'name': 'ListingType', 'value': 'StoreInventory'}
        ],
        'aspectFilter': [
            {'aspectName': 'Release Year', 'aspectValueName': ['2018', '2017']},
            {'aspectName': 'Platform', 'aspectValueName': 'Nintendo Switch'}
        ],
        'paginationInput': {
            'entriesPerPage': '100',
            'pageNumber': str(c)
        },

        'sortOrder': 'StartTimeNewest',
        # 'sortOrder': 'CurrentPriceHighest',
        'outputSelector': ['CategoryHistogram',
                           'AspectHistogram']
    })
    dictstr2 = api.response.dict()

    for item in dictstr2['searchResult']['item']:
        # print(num)
        ws.cell(row=num, column=1, value='{}'.format(item['title']))
        ws.cell(row=num, column=2, value=float('{}'.format(item['sellingStatus']['currentPrice']['value'])))
        ws.cell(row=num, column=3, value='Nintendo')
        ws.cell(row=num, column=4, value='{}'.format(item['primaryCategory']['categoryName']))
        if 'productId' in item:
            ws.cell(row=num, column=5, value=float('{}'.format(item['productId']['value'])))
        ws.cell(row=num, column=6, value='{}'.format(item['viewItemURL']))
        if 'galleryURL' in item:
            ws.cell(row=num, column=7, value='{}'.format(item['galleryURL']))
        ws.cell(row=num, column=8, value='{}'.format(item['itemId']))
        num += 1

wb.save(file)
# print(hist)
# num = 1
# for item in dictstr['aspectHistogramContainer']['aspect']:
#     print(num)
#     print(item)
#     num += 1




